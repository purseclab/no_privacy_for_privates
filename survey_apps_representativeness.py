import argparse
import math
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

list_foreign_developers = [
    'asvab.exam',
    'pp.asvab',
    'asvab.test.prep.us',
]

list_third_party = [
    'Huawei Mobile Services (HMS) Core',
    'Mintegral',
    'Pangle',
    'RjFun',
    'AppMetrica',
    'Pushwoosh',
    'Yandex Ad',
]

problem_pii = [
    'Address',
    'Income',
    'Credit Score',
    'Net Worth',
    'Phone',
    'Birth Date',
]

problem_perms = [
    'Access Camera',
    'Access Approximate Location',
    'Access Fine Location',
    'Access Background Location',
    'Record Audio',
    'Access Contacts',
    'Access SMS / iMessages',
]

PII_APP_ALIASES = {
    'USAA Mobile': 'USAA',
    'ASVAB Flashcards': 'ASVAB Practice Test Prep 2024',
    'ASVAB Mastery: ASVAB Test': 'ASVAB Test 2025 prep',
    'ASVAB Practice for Dummies': 'ASVAB Practice Test Prep 2024',
    'ASVAB Practice Test 2024 Prep': 'ASVAB Practice Test Prep 2024',
    'ATAK-CIV (Civil Use)': 'ATAK',
    'Ace PDG+SJT 2026 USAF PFE Prep': 'ACE PDF Airforce',
    'Ace PDG - Air Force': 'ACE PDF Airforce',
    'Hots&Cots': 'Hots and Cots',
}

PERMISSION_NAME_MAP = {
    'CAMERA': 'Access Camera',
    'ACCESS_COARSE_LOCATION': 'Access Approximate Location',
    'ACCESS_FINE_LOCATION': 'Access Fine Location',
    'ACCESS_BACKGROUND_LOCATION': 'Access Background Location',
    'RECORD_AUDIO': 'Record Audio',
    'READ_CONTACTS': 'Access Contacts',
    'WRITE_CONTACTS': 'Access Contacts',
    'GET_ACCOUNTS': 'Access Contacts',
    'READ_SMS': 'Access SMS / iMessages',
    'RECEIVE_SMS': 'Access SMS / iMessages',
    'SEND_SMS': 'Access SMS / iMessages',
    'READ_PHONE_STATE': 'Make / Manage Phone Calls',
    'CALL_PHONE': 'Make / Manage Phone Calls',
    'PROCESS_OUTGOING_CALLS': 'Make / Manage Phone Calls',
}

def normalize(value: str) -> str:
    if pd.isna(value):
        return ''
    return re.sub(r'[^a-z0-9]+', '', str(value).lower().strip())

def get_pii_row(app_name: str, pii_df: pd.DataFrame):
    candidates = [app_name, PII_APP_ALIASES.get(app_name)]
    for candidate in candidates:
        if not candidate:
            continue
        match = pii_df[pii_df['Smartphone App'].map(normalize) == normalize(candidate)]
        if not match.empty:
            return match.iloc[0]
    return None

def get_permission_row(app_name: str, perms_df: pd.DataFrame):
    match = perms_df[perms_df['App'].map(normalize) == normalize(app_name)]
    if not match.empty:
        return match.iloc[0]
    return None

def get_tpl_row(package_name: str, tpls_df: pd.DataFrame):
    match = tpls_df[tpls_df['App'].map(normalize) == normalize(package_name)]
    if not match.empty:
        return match.iloc[0]
    return None

def percentage(numerator: int, denominator: int) -> float:
    if denominator == 0:
        return 0.0
    return round((numerator / denominator) * 100, 2)

def proportion_ci_pct(x: int, n: int):
    if n == 0:
        return (None, None)
    p = x / n
    se = math.sqrt(max(p * (1 - p) / n, 0))
    lo = max(0.0, p - 1.96 * se) * 100
    hi = min(1.0, p + 1.96 * se) * 100
    return (round(lo, 2), round(hi, 2))

def two_proportion_pvalue(x1, n1, x2, n2):
    if n1 == 0 or n2 == 0:
        return None
    p1 = x1 / n1
    p2 = x2 / n2
    pooled = (x1 + x2) / (n1 + n2)
    variance = pooled * (1 - pooled) * ((1 / n1) + (1 / n2))
    if variance == 0:
        return 1.0
    z = (p1 - p2) / math.sqrt(variance)
    cdf = 0.5 * (1 + math.erf(abs(z) / math.sqrt(2)))
    return round(2 * (1 - cdf), 4)

def mean_ci(values):
    values = [float(v) for v in values]
    n = len(values)
    if n == 0:
        return (None, None)
    mean_v = sum(values) / n
    if n == 1:
        return (round(mean_v, 3), round(mean_v, 3))
    var = sum((v - mean_v) ** 2 for v in values) / (n - 1)
    se = math.sqrt(var / n)
    lo = mean_v - 1.96 * se
    hi = mean_v + 1.96 * se
    return (round(lo, 3), round(hi, 3))

def cohen_d(sample_values, population_values):
    x = [float(v) for v in sample_values]
    y = [float(v) for v in population_values]
    if len(x) < 2 or len(y) < 2:
        return None
    mx = sum(x) / len(x)
    my = sum(y) / len(y)
    vx = sum((v - mx) ** 2 for v in x) / (len(x) - 1)
    vy = sum((v - my) ** 2 for v in y) / (len(y) - 1)
    pooled_num = ((len(x) - 1) * vx) + ((len(y) - 1) * vy)
    pooled_den = len(x) + len(y) - 2
    if pooled_den <= 0:
        return None
    pooled_sd = math.sqrt(pooled_num / pooled_den) if pooled_num > 0 else 0
    if pooled_sd == 0:
        return 0.0
    return round((mx - my) / pooled_sd, 3)

def interpret_ratio(ratio):
    if ratio is None:
        return 'Insufficient data'
    if ratio >= 1.2:
        return 'Survey apps are more problematic'
    if ratio <= 0.8:
        return 'Survey apps are less problematic'
    return 'Survey apps are broadly similar'

def format_sheet(ws):
    header_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.freeze_panes = 'A2'
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            val = '' if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[letter].width = min(max_len + 2, 42)

def build_app_flags(apps_df, pii_df, perms_df, tpls_df):
    pii_rows = []
    perm_rows = []
    tpl_rows = []
    normalized_permissions = []

    for _, row in apps_df.iterrows():
        pii_row = get_pii_row(row['App name'], pii_df)
        perm_row = get_permission_row(row['App name'], perms_df)
        tpl_row = get_tpl_row(row['Package name'], tpls_df)
        pii_rows.append(pii_row)
        perm_rows.append(perm_row)
        tpl_rows.append(tpl_row)

        perms_set = set()
        if perm_row is not None:
            raw_permissions = str(perm_row.get('Permissions', ''))
            for token in raw_permissions.split(','):
                token = token.strip()
                if token:
                    perms_set.add(PERMISSION_NAME_MAP.get(token, token))
        normalized_permissions.append(perms_set)

    out = apps_df.copy()
    out['_pii_row'] = pii_rows
    out['_perm_row'] = perm_rows
    out['_tpl_row'] = tpl_rows
    out['_perm_set'] = normalized_permissions
    out['pop_bucket'] = out['Popularity level'].astype(str).str.strip().str.lower()

    for pii in problem_pii:
        out[f'PII::{pii}'] = out['_pii_row'].apply(
            lambda row, pii=pii: False if row is None else str(row.get(pii, '')).strip().upper() == 'Y'
        )
    for perm in problem_perms:
        out[f'PERMISSION::{perm}'] = out['_perm_set'].apply(lambda permset, perm=perm: perm in permset)
    for package_name in list_foreign_developers:
        out[f'FOREIGN_DEV::{package_name}'] = out['Package name'].astype(str) == package_name
    for tpl in list_third_party:
        out[f'TPL::{tpl}'] = out['_tpl_row'].apply(
            lambda row, tpl=tpl: False if row is None else tpl.lower() in str(row.get('TPLs', '')).lower()
        )

    out['problem_foreign_developers'] = out[[f'FOREIGN_DEV::{x}' for x in list_foreign_developers]].any(axis=1)
    out['problem_third_party'] = out[[f'TPL::{x}' for x in list_third_party]].any(axis=1)
    out['problem_pii'] = out[[f'PII::{x}' for x in problem_pii]].any(axis=1)
    out['problem_perms'] = out[[f'PERMISSION::{x}' for x in problem_perms]].any(axis=1)

    out['count_problem_pii'] = out[[f'PII::{x}' for x in problem_pii]].sum(axis=1)
    out['count_problem_perms'] = out[[f'PERMISSION::{x}' for x in problem_perms]].sum(axis=1)
    out['count_problem_third_party'] = out[[f'TPL::{x}' for x in list_third_party]].sum(axis=1)
    out['count_problem_foreign_developers'] = out[[f'FOREIGN_DEV::{x}' for x in list_foreign_developers]].sum(axis=1)

    return out

def summary_chart_all_apps(apps_df):
    rows = []
    pop_buckets = ['popular', 'mid popular', 'less popular']
    col_map = {
        'popular': '% of Popular Apps',
        'mid popular': '% of Mid Popular Apps',
        'less popular': '% of Less Popular Apps',
    }
    denominators = {b: int((apps_df['pop_bucket'] == b).sum()) for b in pop_buckets}
    mapping = [
        ('list_foreign_developers', 'problem_foreign_developers'),
        ('list_third_party', 'problem_third_party'),
        ('problem_pii', 'problem_pii'),
        ('problem_perms', 'problem_perms'),
    ]
    for label, flag_col in mapping:
        row = {'Problem': label}
        for bucket in pop_buckets:
            bucket_df = apps_df[apps_df['pop_bucket'] == bucket]
            row[col_map[bucket]] = percentage(int(bucket_df[flag_col].sum()), denominators[bucket])
        rows.append(row)
    return pd.DataFrame(rows)

def popularity_counts(apps_df):
    counts = apps_df['pop_bucket'].value_counts()
    return pd.DataFrame([{
        'Popular': int(counts.get('popular', 0)),
        'Mid popular': int(counts.get('mid popular', 0)),
        'Less popular': int(counts.get('less popular', 0)),
        'Total': int(len(apps_df)),
        'Survey apps': int((apps_df['Used in survey'].astype(str).str.strip().str.lower() == 'yes').sum()),
    }])

def representativeness_summary(apps_df):
    survey_df = apps_df[apps_df['Used in survey'].astype(str).str.strip().str.lower() == 'yes'].copy()
    all_df = apps_df.copy()

    categories = [
        ('PII', 'problem_pii', 'count_problem_pii', len(problem_pii)),
        ('Permissions', 'problem_perms', 'count_problem_perms', len(problem_perms)),
        ('Third-party code', 'problem_third_party', 'count_problem_third_party', len(list_third_party)),
        ('Foreign developer', 'problem_foreign_developers', 'count_problem_foreign_developers', len(list_foreign_developers)),
    ]

    rows = []
    for label, flag_col, count_col, n_items in categories:
        x_all = int(all_df[flag_col].sum())
        x_survey = int(survey_df[flag_col].sum())
        n_all = len(all_df)
        n_survey = len(survey_df)
        all_pct = percentage(x_all, n_all)
        survey_pct = percentage(x_survey, n_survey)
        all_ci = proportion_ci_pct(x_all, n_all)
        survey_ci = proportion_ci_pct(x_survey, n_survey)

        mean_all = round(float(all_df[count_col].mean()), 3)
        mean_survey = round(float(survey_df[count_col].mean()), 3)
        mean_all_ci = mean_ci(all_df[count_col].tolist())
        mean_survey_ci = mean_ci(survey_df[count_col].tolist())

        max_share_all = round((mean_all / n_items) * 100, 2)
        max_share_survey = round((mean_survey / n_items) * 100, 2)

        prevalence_ratio = round((survey_pct / all_pct), 3) if all_pct > 0 else None
        burden_ratio = round((mean_survey / mean_all), 3) if mean_all > 0 else None
        prevalence_diff_pp = round(survey_pct - all_pct, 2)
        burden_diff = round(mean_survey - mean_all, 3)
        pvalue = two_proportion_pvalue(x_survey, n_survey, x_all, n_all)
        effect_d = cohen_d(survey_df[count_col].tolist(), all_df[count_col].tolist())

        rows.append({
            'Problem family': label,
            'Survey apps n': n_survey,
            'All apps n': n_all,
            'Survey prevalence %': survey_pct,
            'Survey prevalence 95% CI': f'{survey_ci[0]} to {survey_ci[1]}',
            'All-app prevalence %': all_pct,
            'All-app prevalence 95% CI': f'{all_ci[0]} to {all_ci[1]}',
            'Prevalence difference (pp)': prevalence_diff_pp,
            'Prevalence ratio (survey/all)': prevalence_ratio,
            'Prevalence p-value': pvalue,
            'Survey mean flagged items/app': mean_survey,
            'Survey mean flagged items 95% CI': f'{mean_survey_ci[0]} to {mean_survey_ci[1]}',
            'All-app mean flagged items/app': mean_all,
            'All-app mean flagged items 95% CI': f'{mean_all_ci[0]} to {mean_all_ci[1]}',
            'Burden difference (items/app)': burden_diff,
            'Burden ratio (survey/all)': burden_ratio,
            'Survey mean % of listed issues triggered': max_share_survey,
            'All-app mean % of listed issues triggered': max_share_all,
            'Effect size (Cohen d)': effect_d,
            'Interpretation': interpret_ratio(prevalence_ratio),
        })
    return pd.DataFrame(rows)

def by_popularity_representativeness(apps_df):
    rows = []
    survey_col = apps_df['Used in survey'].astype(str).str.strip().str.lower() == 'yes'
    categories = [
        ('PII', 'problem_pii'),
        ('Permissions', 'problem_perms'),
        ('Third-party code', 'problem_third_party'),
        ('Foreign developer', 'problem_foreign_developers'),
    ]
    for bucket in ['popular', 'mid popular', 'less popular']:
        bucket_df = apps_df[apps_df['pop_bucket'] == bucket].copy()
        survey_df = bucket_df[survey_col.loc[bucket_df.index]]
        nonsurvey_df = bucket_df[~survey_col.loc[bucket_df.index]]
        for label, flag_col in categories:
            survey_n = len(survey_df)
            other_n = len(nonsurvey_df)
            survey_count = int(survey_df[flag_col].sum()) if survey_n else 0
            other_count = int(nonsurvey_df[flag_col].sum()) if other_n else 0
            survey_pct = percentage(survey_count, survey_n)
            other_pct = percentage(other_count, other_n)
            rows.append({
                'Popularity bucket': bucket,
                'Problem family': label,
                'Survey apps in bucket': survey_n,
                'Non-survey apps in bucket': other_n,
                'Survey prevalence %': survey_pct,
                'Non-survey prevalence %': other_pct,
                'Difference (survey - non-survey, pp)': round(survey_pct - other_pct, 2),
                'Ratio (survey/non-survey)': round((survey_pct / other_pct), 3) if other_pct > 0 else None,
                'p-value': two_proportion_pvalue(survey_count, survey_n, other_count, other_n) if survey_n and other_n else None,
            })
    return pd.DataFrame(rows)

def item_breakdown(apps_df, items, prefix, label_name):
    rows = []
    for item in items:
        survey_mask = apps_df['Used in survey'].astype(str).str.strip().str.lower() == 'yes'
        all_count = int(apps_df[f'{prefix}::{item}'].sum())
        survey_count = int(apps_df.loc[survey_mask, f'{prefix}::{item}'].sum())
        all_pct = percentage(all_count, len(apps_df))
        survey_pct = percentage(survey_count, int(survey_mask.sum()))
        rows.append({
            label_name: item,
            'All apps count': all_count,
            'All apps %': all_pct,
            'Survey apps count': survey_count,
            'Survey apps %': survey_pct,
            'Difference (pp)': round(survey_pct - all_pct, 2),
            'Ratio (survey/all)': round((survey_pct / all_pct), 3) if all_pct > 0 else None,
        })
    return pd.DataFrame(rows)

def create_debug_sheet(apps_df):
    cols = [
        'App name', 'Package name', 'Download count', 'Popularity level', 'Used in survey',
        'problem_foreign_developers', 'problem_third_party', 'problem_pii', 'problem_perms',
        'count_problem_foreign_developers', 'count_problem_third_party',
        'count_problem_pii', 'count_problem_perms'
    ]
    cols += [f'FOREIGN_DEV::{x}' for x in list_foreign_developers]
    cols += [f'TPL::{x}' for x in list_third_party]
    cols += [f'PII::{x}' for x in problem_pii]
    cols += [f'PERMISSION::{x}' for x in problem_perms]
    return apps_df[cols].copy()

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--target-sheet', default='target_sheet.xlsx')
    parser.add_argument('--pii-csv', default='pii_collection.csv')
    parser.add_argument('--permissions-csv', default='permission_collection_named.csv')
    parser.add_argument('--tpls-csv', default='tpls.csv')
    parser.add_argument('--output', default='problem_chart_all_apps_stronger.xlsx')
    args = parser.parse_args()

    apps_df = pd.read_excel(args.target_sheet)
    pii_df = pd.read_csv(args.pii_csv)
    perms_df = pd.read_csv(args.permissions_csv)
    tpls_df = pd.read_csv(args.tpls_csv)

    flagged = build_app_flags(apps_df, pii_df, perms_df, tpls_df)

    with pd.ExcelWriter(args.output, engine='openpyxl') as writer:
        summary_chart_all_apps(flagged).to_excel(writer, sheet_name='problem_chart_all_apps', index=False)
        popularity_counts(flagged).to_excel(writer, sheet_name='popularity_counts', index=False)
        representativeness_summary(flagged).to_excel(writer, sheet_name='representativeness_strong', index=False)
        by_popularity_representativeness(flagged).to_excel(writer, sheet_name='repr_by_popularity', index=False)
        item_breakdown(flagged, problem_pii, 'PII', 'PII').to_excel(writer, sheet_name='pii_breakdown', index=False)
        item_breakdown(flagged, problem_perms, 'PERMISSION', 'Permission').to_excel(writer, sheet_name='permission_breakdown', index=False)
        item_breakdown(flagged, list_foreign_developers, 'FOREIGN_DEV', 'Foreign developer').to_excel(writer, sheet_name='foreign_dev_breakdown', index=False)
        item_breakdown(flagged, list_third_party, 'TPL', 'Third party').to_excel(writer, sheet_name='third_party_breakdown', index=False)
        create_debug_sheet(flagged).to_excel(writer, sheet_name='all_apps_debug', index=False)

    wb = load_workbook(args.output)
    for ws in wb.worksheets:
        format_sheet(ws)
    wb.save(args.output)

if __name__ == '__main__':
    main()
